"""爬取 hfutxc 寝室卫生检查系统——查询宿舍床铺评分的数据"""
import os
import time
import csv
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule

from dorm_dict import DORM_DICT

URL = "http://39.106.82.121/query/getStudentScore"  # 请求地址
GET = False  # 是否请求，True：请求数据并保存然后处理，False：读取保存的数据然后处理
DELAY = 0.1  # 每次请求间隔时间，单位秒

# 寝室楼栋，范围如下：1N, 1S, 2N, 2S, 3S, 4N, 4S, 5N, 5S, 6N, 6S, 7#, 8#, 9N, 9S, 10N, 10S
BUILDING = "9N"

YEAR_TERM_INDEX = ((23, 1),)  # 目标学期，格式为(年,学期序号)，单个学期需在tuple后打,
WEEK_NUM = 20  # 学期的周数，默认为20


def year_term_get(date: str) -> tuple:
    """使用日期计算学期"""
    year = int(date[2:4])
    month = int(date[5:7])
    if month <= 2:
        year -= 1
        term = 1
    elif month <= 8:
        year -= 1
        term = 2
    else:
        term = 1
    return (year, term)


def dorm_req(dorm: str) -> bool:
    """请求并保存获取的数据表格为htm文件，返回请求是否有效"""
    try:
        response = requests.get(URL, params={"student_code": dorm}, timeout=10)
        response.encoding = "UTF-8"
        # http状态码异常
        if response.status_code != 200:
            print(f"Request {dorm} failed! Code: {response.status_code}")
            return False
        text = response.text
        # 无数据则不保存
        if text == '<tr><td colspan="5" align="center">无数据</td></tr>\n':
            return False
        with open(f"{BUILDING}\\{dorm}.htm", 'w+', encoding='UTF-8') as f:
            f.write(text)
        return True
    except requests.exceptions.Timeout:  # 超时
        print(f"Request {dorm} timed out!")
        return False
    except requests.RequestException as e:  # 请求错误
        print(f"Request {dorm} failed! Error: {e}")
        return False


def dorm_dec(dorm: str) -> list:
    """解码保存的指定寝室的数据，返回指定学期内的成绩"""
    date_index = ["-1"]*WEEK_NUM*len(YEAR_TERM_INDEX)
    try:
        with open(f"{BUILDING}\\{dorm}.htm", 'r', encoding='UTF-8') as f:
            html_content = f.read()
    except FileNotFoundError:
        return date_index
    soup = BeautifulSoup(html_content, 'html.parser')
    rows = soup.find_all('tr')
    for row in rows:
        cols = row.find_all('td')
        # 排除一次成绩（四行）中的非首行，其宽1，首行宽5
        if len(cols) != 5:
            continue
        # 处理日期填入位置错误情况
        if cols[4].text != "--":
            date = cols[4].text
        else:
            date = cols[2].text[:10]
        # 判断是否为目标学期
        term = year_term_get(date)
        if term in YEAR_TERM_INDEX:
            week = int(cols[0].text)-1
            # 处理周数填入数值错误情况
            if week >= WEEK_NUM:
                continue
            pos = week+WEEK_NUM*YEAR_TERM_INDEX.index(term)
            date_index[pos] = cols[1].text
    return date_index


def remove_empty_weeks(table: list) -> list:
    """清理无数据周"""
    columns_to_delete = []
    # 统计全无数据周
    for col_index in range(1, len(table[0])):
        if all(row[col_index] == "-1" for row in table[1:]):
            columns_to_delete.append(col_index)
    # 清理无数据周
    for row in table:
        for col_index in reversed(columns_to_delete):
            del row[col_index]

    return table


def col_to_excel(number: int) -> str:
    """将整数转换为excel中列的字母序号"""
    result = ""
    while number > 0:
        remainder = (number - 1) % 26
        result = chr(ord('A') + remainder) + result
        number = (number - 1) // 26

    return result


if __name__ == '__main__':
    # 创建数据存放路径
    if not os.path.exists(BUILDING):
        os.makedirs(BUILDING)
        GET = True

    # 生成表头
    head = ["寝室"]
    for column in range(len(YEAR_TERM_INDEX)):
        for num in range(WEEK_NUM):
            head.append(str(num+1))
    head.append("平均成绩")
    output = [head]

    # 处理数据
    for dorm_name in DORM_DICT[BUILDING]:
        if GET:
            time.sleep(DELAY)
            if not dorm_req(dorm_name):  # 这一步会请求成绩
                continue
        score = dorm_dec(dorm_name)
        # 跳过空寝室
        if all(item == "-1" for item in score):
            continue
        # 计算有效平均成绩
        score_int = [int(num) for num in score if int(num) != -1]
        average = sum(score_int) / len(score_int)
        # 添加首尾列
        score.insert(0, dorm_name)
        score.append(f"{average:.2f}")
        print(",".join(score))
        output.append(score)

    # 清理无数据周
    output = remove_empty_weeks(output)
    WEEK_COUNT = len(output[0])-2
    DORM_COUNT = len(output)-1

    # 从旧到新按照成绩排序
    head = output.pop(0)
    for i in range(0, WEEK_COUNT+1):
        output.sort(key=lambda x, i=i: float(x[i+1]), reverse=True)
    output.insert(0, head)

    # 添加序号列
    for i, output_row in enumerate(output[1:], start=1):
        output_row.insert(0, i)
    output[0].insert(0, "序号")

    # 生成输出文件名
    output_filename = f"{BUILDING}-{DORM_COUNT}-"
    for year_term in YEAR_TERM_INDEX:
        for item in year_term:
            output_filename += f"{str(item)}-"
    output_filename += str(WEEK_COUNT)

    # 保存csv
    with open(f"{output_filename}.csv", 'w', newline='', encoding='GBK') as output_file:
        writer = csv.writer(output_file)
        writer.writerows(output)

    # 创建Excel表格
    wb = Workbook()
    ws = wb.active
    # 将数据写入Excel文件
    for row_index, row_data in enumerate(output, start=1):
        for col_index, value in enumerate(row_data, start=1):
            try:
                # 尝试将值转换为整数（成绩值）
                cell_value = int(value)
            except ValueError:
                try:
                    # 尝试将值转换为浮点数（平均成绩值）
                    cell_value = float(value)
                except ValueError:
                    # 如果都失败，保留原始值
                    cell_value = value
            ws.cell(row=row_index, column=col_index, value=cell_value)
    # 给全表添加字体和单元格对齐属性
    for row in ws[f"A1:{col_to_excel(WEEK_COUNT+3)}{DORM_COUNT+1}"]:
        for cell in row:
            cell.font = Font(name='微软雅黑', size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    # 给成绩值添加0的数字格式
    for row in ws[f"C2:{col_to_excel(WEEK_COUNT+2)}{DORM_COUNT+1}"]:
        for cell in row:
            cell.number_format = '0'
    # 给平均成绩值添加0.00的数字格式
    for row in ws[f"{col_to_excel(WEEK_COUNT+3)}2:{col_to_excel(WEEK_COUNT+3)}{DORM_COUNT+1}"]:
        for cell in row:
            cell.number_format = '0.00'
    # 给所有成绩值添加蓝-白-红的条件格式
    rule = ColorScaleRule(start_type='percentile', start_value=0, start_color='F8696B',
                          mid_type='percentile', mid_value=50, mid_color='FCFCFF',
                          end_type='percentile', end_value=100, end_color='5A8AC6')
    ws.conditional_formatting.add(
        f"C2:{col_to_excel(WEEK_COUNT+3)}{DORM_COUNT+1}", rule)
    # 保存Excel表格
    wb.save(f"{output_filename}.xlsx")
    wb.close()

    print(f"Done! #validate dorm:{DORM_COUNT} week:{WEEK_COUNT}")
