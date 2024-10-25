"""根据存在的数据文件，统计寝室存在情况，生成寝室字典dorm_dict.py"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from dorm_decode import col_to_excel

# 寝室楼栋
BUILDINGS = ["1N", "1S", "2N", "2S", "3N", "3S", "4N", "4S", "5N",
             "5S", "6N", "6S", "7#", "8#", "9N", "9S", "10N", "10S"]

# 院系
FACULTIES = ["生态", "文法", "英语", "城市", "电气", "计算",
             "经济", "能源", "食品", "物流", "材料", "机械"]

# 年级
GRADES = ["21", "22", "23", "24"]

# 遍历层号范围，默认为range(1, 7)
FLOORS = range(1, 7)

# 遍历房间号范围，默认为range(1, 41)
ROOMS = range(1, 41)  # 房间号范围，默认为range(1, 41)

DICT_XLSX_FACULTY = True


def dorm_dict_genf(buildings: list, floors: list, rooms: list, dict_xlsx_faculty: bool) -> None:
    """生成楼栋字典与院系年级字典"""
    # 创建空的楼栋字典
    dorm_dict = {key: [] for key in buildings}

    # 创建空的院系年级字典
    faculty_grade = []
    for faculty in FACULTIES:
        for grade in GRADES:
            faculty_grade.append(faculty+grade)
    faculty_dict = {key: [] for key in faculty_grade}

    # 创建Excel表格
    wb = Workbook()
    ws = wb.active

    # 生成首行并创建数据存放路径
    for building_index, building in enumerate(buildings):
        if not os.path.exists(building):
            raise FileNotFoundError(f"path \\{building}\\ not fonud")
        col_index = building_index+2
        ws.cell(row=1, column=col_index, value=building)

    # 生成数据
    for floor_index, floor in enumerate(floors, start=1):
        for room_index, room in enumerate(rooms, start=1):
            # 生成首列
            row_index = (floor_index-1)*len(rooms)+room_index+1
            ws.cell(row=row_index, column=1, value=int(f"{floor}{room:02d}"))
            # 循环楼栋查找寝室
            for building_index, building in enumerate(buildings):
                col_index = building_index+2
                dorm_name = f"{building}{floor}{room:02d}"
                # 查找是否存在对应文件
                if os.path.exists(f"{building}\\{dorm_name}.htm"):
                    # 在楼栋字典中添加对应寝室
                    dorm_dict[building].append(dorm_name)
                    with open(f"{building}\\{dorm_name}.htm", 'r', encoding='UTF-8') as f:
                        html_content = f.read()
                    # 查找系
                    pos = int(html_content.find("faculty_id = "))
                    faculty_id = int(html_content[pos+13:pos+15])
                    # 查找年级
                    pos = int(html_content.find("grade_id = "))
                    grade_id = int(html_content[pos+11:pos+12])
                    string = FACULTIES[faculty_id-1]+GRADES[grade_id-6]
                    ws.cell(row=row_index, column=col_index, value=string)
                    # 在院系年级字典中添加对应寝室
                    faculty_dict[string].append(dorm_name)

    # 保存字典
    with open("dorm_dict.py", 'w+', encoding='UTF-8') as f:
        f.write(f"BUILDING_DICT = {str(dorm_dict)}\n")
        f.write(f"FACULTY_DICT = {str(faculty_dict)}\n")

    if dict_xlsx_faculty:
        # 给全表添加字体和单元格对齐属性
        for row in ws[f"A1:{col_to_excel(len(buildings)+1)}{len(floors)*len(rooms)+1}"]:
            for cell in row:
                cell.font = Font(name='微软雅黑', size=12)
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')

        # 保存Excel表格
        wb.save("dorm_dict.xlsx")
        wb.close()
        print("dorm_dict.xlsx & dorm_dict.py Generated!")
    else:
        print("dorm_dict.py Generated!")


if __name__ == "__main__":
    dorm_dict_genf(BUILDINGS, FLOORS, ROOMS, DICT_XLSX_FACULTY)
