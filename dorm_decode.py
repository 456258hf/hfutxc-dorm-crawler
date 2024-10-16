"""爬取 hfutxc 寝室卫生检查系统——查询宿舍床铺评分的数据"""
import csv
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule

from config import WEEK_NUM


def year_term_get(date: str) -> tuple:
    """使用日期计算学期"""
    year = int(date[2:4])
    month = int(date[5:7])
    if month <= 2:
        year -= 1
        term = 1
    elif month <= 8:
        year -= 1
        term = 2
    else:
        term = 1
    return (year, term)


def dorm_dec(dorm: str, year_term_index: list) -> list:
    """解码保存的指定寝室的数据，返回指定学期内的成绩"""
    date_index = ["-1"]*WEEK_NUM*len(year_term_index)
    # 匹配楼栋号，读取文件
    building = dorm[:-3]
    try:
        with open(f"{building}\\{dorm}.htm", 'r', encoding='UTF-8') as f:
            html_content = f.read()
    except FileNotFoundError:
        return date_index
    # 在调试信息前截断html_content
    point = html_content.find('<div')
    html_content = html_content[:point]
    # 使用bs4处理html
    soup = BeautifulSoup(html_content, 'html.parser')
    rows = soup.find_all('tr')
    for row in rows:
        cols = row.find_all('td')
        # 排除一次成绩（四行）中的非首行，其宽1，首行宽6
        if len(cols) != 6:
            continue
        # 处理日期填入位置错误情况
        if cols[4].text != "--":
            date = cols[4].text
        else:
            date = cols[2].text[:10]
        # 判断是否为目标学期
        term = year_term_get(date)
        if term in year_term_index:
            week = int(cols[0].text)-1
            # 处理周数填入数值错误情况
            if week >= WEEK_NUM:
                continue
            pos = week+WEEK_NUM*year_term_index.index(term)
            date_index[pos] = cols[1].text
    return date_index


def remove_empty_weeks(table: list) -> list:
    """清理无数据周"""
    columns_to_delete = []
    # 统计全无数据周
    for col_index in range(1, len(table[0])):
        if all(row[col_index] == "-1" for row in table[1:]):
            columns_to_delete.append(col_index)
    # 清理无数据周
    for row in table:
        for col_index in reversed(columns_to_delete):
            del row[col_index]

    return table


def col_to_excel(number: int) -> str:
    """将整数转换为excel中列的字母序号"""
    result = ""
    while number > 0:
        remainder = (number - 1) % 26
        result = chr(ord('A') + remainder) + result
        number = (number - 1) // 26

    return result


def dorms_dec(group: str, dorms: list, year_term_index: list, extension: list) -> None:
    """将指定的寝室们的指定学期的数据进行处理，生成指定格式的文件"""
    # 生成表头
    head = ["寝室"]
    for num in range(WEEK_NUM*len(year_term_index)):
        head.append(str((num % WEEK_NUM)+1))
    head.append("平均成绩")
    output = [head]

    # 处理数据
    for dorm_name in dorms:
        score = dorm_dec(dorm_name, year_term_index)
        # 跳过空寝室
        if all(item == "-1" for item in score):
            continue
        # 计算有效平均成绩
        score_int = [int(num) for num in score if int(num) != -1]
        average = sum(score_int) / len(score_int)
        # 添加首尾列
        score.insert(0, dorm_name)
        score.append(f"{average:.2f}")
        output.append(score)

    # 清理无数据周
    output = remove_empty_weeks(output)
    week_count = len(output[0])-2
    dorm_count = len(output)-1

    # 从旧到新按照成绩排序
    head = output.pop(0)
    for i in range(0, week_count+1):
        output.sort(key=lambda x, i=i: float(x[i+1]), reverse=True)
    output.insert(0, head)

    # 添加序号列
    for i, output_row in enumerate(output[1:], start=1):
        output_row.insert(0, i)
    output[0].insert(0, "序号")

    # 生成输出文件名
    output_filename = f"{group}-{dorm_count}-"
    for year_term in year_term_index:
        for item in year_term:
            output_filename += f"{str(item)}-"
    output_filename += str(week_count)

    if "csv" in extension:
        # 保存csv
        with open(f"{output_filename}.csv", 'w', newline='', encoding='GBK') as output_file:
            writer = csv.writer(output_file)
            writer.writerows(output)

    if "xlsx" in extension:
        # 创建Excel表格
        wb = Workbook()
        ws = wb.active
        # 将数据写入Excel文件
        for row_index, row_data in enumerate(output, start=1):
            for col_index, value in enumerate(row_data, start=1):
                try:
                    # 尝试将值转换为整数（成绩值）
                    cell_value = int(value)
                except ValueError:
                    try:
                        # 尝试将值转换为浮点数（平均成绩值）
                        cell_value = float(value)
                    except ValueError:
                        # 如果都失败，保留原始值
                        cell_value = value
                ws.cell(row=row_index, column=col_index, value=cell_value)
        # 给全表添加字体和单元格对齐属性
        for row in ws[f"A1:{col_to_excel(week_count+3)}{dorm_count+1}"]:
            for cell in row:
                cell.font = Font(name='微软雅黑', size=12)
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
        # 给成绩值添加0的数字格式
        for row in ws[f"C2:{col_to_excel(week_count+2)}{dorm_count+1}"]:
            for cell in row:
                cell.number_format = '0'
        # 给平均成绩值添加0.00的数字格式
        for row in ws[f"{col_to_excel(week_count+3)}2:{col_to_excel(week_count+3)}{dorm_count+1}"]:
            for cell in row:
                cell.number_format = '0.00'
        # 给所有成绩值添加蓝-白-红的条件格式
        rule = ColorScaleRule(start_type='percentile', start_value=0, start_color='F8696B',
                              mid_type='percentile', mid_value=50, mid_color='FCFCFF',
                              end_type='percentile', end_value=100, end_color='5A8AC6')
        ws.conditional_formatting.add(
            f"C2:{col_to_excel(week_count+3)}{dorm_count+1}", rule)
        # 保存Excel表格
        wb.save(f"{output_filename}.xlsx")
        wb.close()

    print(f"{group} done! #validate dorm:{dorm_count} week:{week_count}")
