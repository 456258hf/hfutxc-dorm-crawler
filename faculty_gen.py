"""根据存在的数据文件，统计寝室存在情况，生成寝室字典dorm_dict.py"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule

from dorm_decode import col_to_excel
from config import BUILDINGS, FACULTIES, GRADES

FLOORS = range(1, 7)  # 层号范围，默认为range(1, 7)
ROOMS = range(1, 41)  # 房间号范围，默认为range(1, 41)
IF_EXCEL = True


def faculty_gen(buildings: list, floors: list, rooms: list, if_excel: bool) -> None:
    """创建空的寝室院系年级字典"""
    faculty_grade=[]
    for faculty in FACULTIES:
        for grade in GRADES:
            faculty_grade.append(faculty+grade)
    faculty_dict = {key: [] for key in faculty_grade}

    # 创建Excel表格
    wb = Workbook()
    ws = wb.active

    # 生成首行并创建数据存放路径
    for building_index, building in enumerate(buildings):
        if not os.path.exists(building):
            raise FileNotFoundError(f"path \\{building}\\ not fonud")
        col_index = building_index+2
        ws.cell(row=1, column=col_index, value=building)

    # 生成数据
    for floor_index, floor in enumerate(floors, start=1):
        for room_index, room in enumerate(rooms, start=1):
            # 生成首列
            row_index = (floor_index-1)*len(rooms)+room_index+1
            ws.cell(row=row_index, column=1, value=f"{floor}{room:02d}")
            # 循环楼栋查找寝室
            for building_index, building in enumerate(buildings):
                col_index = building_index+2
                dorm_name = f"{building}{floor}{room:02d}"
                # 查找对应寝室年级与系
                if os.path.exists(f"{building}\\{dorm_name}.htm"):
                    with open(f"{building}\\{dorm_name}.htm", 'r', encoding='UTF-8') as f:
                        html_content = f.read()
                    # 查找系
                    pos = int(html_content.find("faculty_id = "))
                    faculty_id = int(html_content[pos+13:pos+15])
                    # 查找年级
                    pos = int(html_content.find("grade_id = "))
                    grade_id = int(html_content[pos+11:pos+12])
                    string = FACULTIES[faculty_id-1]+GRADES[grade_id-6]
                    ws.cell(row=row_index, column=col_index, value=string)
                    faculty_dict[string].append(dorm_name)

    # 给全表添加字体和单元格对齐属性
    for row in ws[f"A1:{col_to_excel(len(buildings)+1)}{len(floors)*len(rooms)+1}"]:
        for cell in row:
            cell.font = Font(name='微软雅黑', size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 添加条件格式
    # rule = ColorScaleRule(start_type='percentile', start_value=0, start_color='F8696B',
    #                       mid_type='percentile', mid_value=50, mid_color='FCFCFF',
    #                       end_type='percentile', end_value=100, end_color='5A8AC6')
    # ws.conditional_formatting.add(
    #     f"B2:{col_to_excel(len(buildings)+1)}{len(floors)*len(rooms)+1}", rule)

    # 保存字典
    with open("faculty_dict.py", 'w+', encoding='UTF-8') as f:
        f.write(f"FACULTY_DICT = {str(faculty_dict)}\n")

    if if_excel:
        # 保存Excel表格
        wb.save("faculty_dict.xlsx")
        wb.close()
        print("faculty_dict.xlsx & faculty_dict.py Generated!")
    else:
        print("faculty_dict.py Generated!")


faculty_gen(BUILDINGS, FLOORS, ROOMS, IF_EXCEL)
